"""Cross-reference scrape.log results against the 88 URLs in version2.xlsx.

Outputs: how many v2 URLs successfully crawled, with notice counts; how many
failed; how many have no scraper at all.

Writes a UTF-8 report to `.scrape-log/v2_results.txt` so Korean text survives
the Windows cp949 console.
"""
import re
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
LOG = ROOT / ".scrape-log" / "scrape.log"
XLSX = ROOT / "version2.xlsx"
OUT = ROOT / ".scrape-log" / "v2_results.txt"

def norm(url: str) -> str:
    if not url:
        return ""
    try:
        p = urlsplit(url.strip())
    except Exception:
        return url.strip()
    path = p.path.rstrip("/") or "/"
    return urlunsplit(("https", p.netloc.lower(), path, p.query, ""))

# 1. Load v2 URLs
v2 = []
wb = load_workbook(XLSX)
for sn in wb.sheetnames:
    ws = wb[sn]
    last_r, last_s = None, None
    for r in range(3, ws.max_row + 1):
        link = ws.cell(row=r, column=5)
        if not link.hyperlink:
            continue
        region = (ws.cell(row=r, column=2).value or "").strip() or last_r or "-"
        sub = (ws.cell(row=r, column=3).value or "").strip() or last_s or "-"
        page = (ws.cell(row=r, column=4).value or "").strip() or "공지사항"
        last_r, last_s = region, sub
        v2.append({"sheet": sn, "region": region, "sub": sub, "page": page,
                   "url": link.hyperlink.target, "url_norm": norm(link.hyperlink.target)})

# 2. Parse scrape.log
log_text = LOG.read_text(encoding="utf-8", errors="ignore")
lines = log_text.splitlines()
header_re = re.compile(r"^── (.+?) / (.+?) / (.+?)$")
url_re = re.compile(r"^\s+(https?://\S+)\s*$")
count_re = re.compile(r"^\s*→ (\d+) notices?\s*$")
fail_re = re.compile(r"^  scrapers\.[\w.]+/.+?  .+$")

results = []
i, n = 0, len(lines)
while i < n:
    m = header_re.match(lines[i])
    if not m:
        i += 1
        continue
    region, sub, page = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
    j = i + 1
    while j < n and not lines[j].strip():
        j += 1
    url = url_re.match(lines[j]).group(1) if j < n and url_re.match(lines[j]) else None
    count = None
    k = j + 1
    while k < n and not header_re.match(lines[k]):
        cm = count_re.match(lines[k])
        if cm:
            count = int(cm.group(1))
        k += 1
    results.append({"region": region, "sub": sub, "page": page, "url": url,
                    "url_norm": norm(url) if url else None, "count": count})
    i = k

crawled = {}
for r in results:
    if not r["url_norm"]:
        continue
    crawled.setdefault(r["url_norm"], 0)
    if r["count"] is not None:
        crawled[r["url_norm"]] += r["count"]

fail_lines = [l for l in lines if fail_re.match(l)]

# 3. Match
matched, unmatched = [], []
for v in v2:
    if v["url_norm"] in crawled:
        matched.append((v, crawled[v["url_norm"]]))
    else:
        unmatched.append(v)
succ = [m for m in matched if m[1] > 0]
zero = [m for m in matched if m[1] == 0]

# 4. Write UTF-8 report
out_lines = []
def p(*args):
    out_lines.append(" ".join(str(a) for a in args))

p("=" * 72)
p(f"  RESULT: {len(succ)}/{len(v2)} v2 sources crawled successfully")
p("=" * 72)
p(f"  - successful (>=1 notice):    {len(succ)}")
p(f"  - matched scraper, 0 notices: {len(zero)}")
p(f"  - no scraper for this URL:    {len(unmatched)}")
p(f"  Total notices upserted to notices_v2 (across all clt-plus scrapers): {sum(r['count'] or 0 for r in results)}")
p(f"  Notices specifically tied to v2 URLs: {sum(c for _, c in matched)}")
p(f"  Sub-source error lines (informational): {len(fail_lines)}")
p("")

p("=== Successful v2 URLs (sorted by notice count) ===")
for v, c in sorted(succ, key=lambda x: (-x[1], x[0]["region"])):
    p(f"  [{c:>4}] {v['region']} / {v['sub']} / {v['page']}")

p("")
p("=== Matched but 0 notices (scraper exists but failed at runtime) ===")
for v, c in zero:
    p(f"  [   0] {v['region']} / {v['sub']} / {v['page']}")
    p(f"          {v['url']}")

p("")
p("=== v2 URLs with NO scraper (need new code) ===")
by_region = {}
for v in unmatched:
    by_region.setdefault(v["region"], []).append(v)
for region, items in sorted(by_region.items()):
    p(f"  -- {region} ({len(items)}):")
    for v in items:
        p(f"       {v['sub']} / {v['page']}")
        p(f"         {v['url']}")

OUT.write_text("\n".join(out_lines), encoding="utf-8")
print(f"Wrote {OUT} ({len(out_lines)} lines)")
print(f"RESULT: {len(succ)}/{len(v2)} v2 sources crawled successfully ({len(zero)} returned 0, {len(unmatched)} have no scraper)")
